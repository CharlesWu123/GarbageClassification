# -*- coding: utf-8 -*-
'''
@Version : 0.1
@Author : Charles
@Time : 2022/11/20 11:08 
@File : io_utils.py 
@Desc : 
'''
import itertools
import os.path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import get_column_letter


def report_confusion_to_excel(report, confusion, target_names, save_path, sheet_names=['report', 'confusion']):
    report.pop('accuracy')
    report_df = pd.DataFrame(list(report.values()), index=list(report.keys()))
    if os.path.exists(save_path):
        with pd.ExcelWriter(save_path, mode='a') as writer:
            report_df.to_excel(writer, sheet_name=sheet_names[0])
    else:
        report_df.to_excel(save_path, sheet_name=sheet_names[0])
    confusion_df = pd.DataFrame(confusion, index=target_names, columns=target_names)
    # 获取 混淆矩阵 错误单元格
    fill_cell = []
    for i in range(len(target_names)):
        for j in range(len(target_names)):
            if i == j: continue
            if confusion[i][j] > 0:
                fill_cell.append('{}{}'.format(get_column_letter(j+2), i+2))
    with pd.ExcelWriter(save_path, mode='a') as writer:
        confusion_df.to_excel(writer, sheet_name=sheet_names[1])
    # 填充颜色
    wb = load_workbook(save_path)
    ws = wb[sheet_names[1]]
    fill = PatternFill(patternType='solid', fgColor='FFFF00')
    for rc in fill_cell:
        ws[rc].fill = fill
    wb.save(save_path)


def plot_confusion_matrix(cmtx, num_classes, class_names=None, figsize=None):
    """
    A function to create a colored and labeled confusion matrix matplotlib figure
    given true labels and preds.
    Args:
        cmtx (ndarray): confusion matrix.
        num_classes (int): total number of classes.
        class_names (Optional[list of strs]): a list of class names.
        figsize (Optional[float, float]): the figure size of the confusion matrix.
            If None, default to [6.4, 4.8].

    Returns:
        img (figure): matplotlib figure.
    """
    if class_names is None or type(class_names) != list:
        class_names = [str(i) for i in range(num_classes)]

    figure = plt.figure(figsize=figsize)
    plt.imshow(cmtx, interpolation="nearest", cmap=plt.cm.Blues)
    plt.title("Confusion matrix")
    plt.colorbar()
    tick_marks = np.arange(len(class_names))
    plt.xticks(tick_marks, class_names, rotation=45)
    plt.yticks(tick_marks, class_names)

    # Use white text if squares are dark; otherwise black.
    threshold = cmtx.max() / 2.0
    for i, j in itertools.product(range(cmtx.shape[0]), range(cmtx.shape[1])):
        color = "white" if cmtx[i, j] > threshold else "black"
        plt.text(
            j,
            i,
            format(cmtx[i, j], ".2f") if cmtx[i, j] != 0 else ".",
            horizontalalignment="center",
            color=color,
        )

    plt.tight_layout()
    plt.ylabel("True label")
    plt.xlabel("Predicted label")

    return figure


def add_confusion_matrix(
        writer,
        cmtx,
        num_classes,
        global_step=None,
        subset_ids=None,
        class_names=None,
        tag="Confusion Matrix",
        figsize=None,
):
    """
    Calculate and plot confusion matrix to a SummaryWriter.
    Args:
        writer (SummaryWriter): the SummaryWriter to write the matrix to.
        cmtx (ndarray): confusion matrix.
        num_classes (int): total number of classes.
        global_step (Optional[int]): current step.
        subset_ids (list of ints): a list of label indices to keep.
        class_names (list of strs, optional): a list of all class names.
        tag (str or list of strs): name(s) of the confusion matrix image.
        figsize (Optional[float, float]): the figure size of the confusion matrix.
            If None, default to [6.4, 4.8].

    """
    if subset_ids is None or len(subset_ids) != 0:
        # If class names are not provided, use class indices as class names.
        if class_names is None:
            class_names = [str(i) for i in range(num_classes)]
        # If subset is not provided, take every classes.
        if subset_ids is None:
            subset_ids = list(range(num_classes))

        sub_cmtx = cmtx[subset_ids, :][:, subset_ids]
        sub_names = [class_names[j] for j in subset_ids]

        sub_cmtx = plot_confusion_matrix(
            sub_cmtx,
            num_classes=len(subset_ids),
            class_names=sub_names,
            figsize=figsize,
        )
        # Add the confusion matrix image to writer.
        writer.add_figure(tag=tag, figure=sub_cmtx, global_step=global_step)



if __name__ == '__main__':
    import numpy as np
    report = {
        '0': {
            'a': 0.1,
            'b': 0.3,
            'c': 0.4
        },
        '1': {
            'a': 0.1,
            'b': 0.3,
            'c': 0.4
        },
        '2': {
            'a': 0.1,
            'b': 0.3,
            'c': 0.4
        }
    }
    confusion = np.random.random((5, 5))
    target_names = ['a', 'b', 'c', 'd', 'e']
    save_path = 'test.xlsx'
    report_confusion_to_excel(report, confusion, target_names, save_path)
    # print(rgb_to_hex(255, 165, 1))